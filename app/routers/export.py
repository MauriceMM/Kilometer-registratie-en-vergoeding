"""
Export routes: genereer een maandoverzicht voor SAP SuccessFactors.
Per dag: datum + totaal km (retour). Exporteerbaar als Excel.
"""

import io
import logging
from datetime import datetime, date
from typing import List
from collections import defaultdict

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.auth import require_login
from app.database import get_db
from app.models import Trip, TripType, TripStatus, Settings

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/export", tags=["export"])

MAANDEN_NL = [
    "", "januari", "februari", "maart", "april", "mei", "juni",
    "juli", "augustus", "september", "oktober", "november", "december"
]


def _haal_instelling(db: Session, key: str, default: str = "") -> str:
    row = db.query(Settings).filter(Settings.key == key).first()
    return row.value if row and row.value else default


def _ritten_voor_maand(db: Session, jaar: int, maand: int) -> List[Trip]:
    """Haal alle afgesloten zakelijke ritten op voor een gegeven maand."""
    start = datetime(jaar, maand, 1)
    if maand == 12:
        eind = datetime(jaar + 1, 1, 1)
    else:
        eind = datetime(jaar, maand + 1, 1)
    return (
        db.query(Trip)
        .filter(
            Trip.status == TripStatus.AFGESLOTEN,
            Trip.type == TripType.ZAKELIJK,
            Trip.datum >= start,
            Trip.datum < eind,
            Trip.afstand_km.isnot(None),
        )
        .order_by(Trip.datum)
        .all()
    )


def _groepeer_per_dag(ritten: List[Trip]) -> dict:
    """
    Groepeer ritten per kalenderdag.
    SAP SuccessFactors verwacht één invoer per dag met het TOTAAL km (retour).
    """
    per_dag = defaultdict(list)
    for rit in ritten:
        dag = rit.datum.date() if rit.datum else date.today()
        per_dag[dag].append(rit)
    return dict(sorted(per_dag.items()))


@router.get("/successfactors/{jaar}/{maand}")
def export_successfactors_overzicht(
    jaar: int,
    maand: int,
    db: Session = Depends(get_db),
    user: str = Depends(require_login),
):
    """
    JSON-overzicht voor SAP SuccessFactors invoer:
    één regel per dag met totaal km en omschrijving.
    """
    ritten = _ritten_voor_maand(db, jaar, maand)
    per_dag = _groepeer_per_dag(ritten)

    resultaat = []
    totaal_km = 0.0
    for dag, dag_ritten in per_dag.items():
        km_dag = sum(r.afstand_km or 0 for r in dag_ritten)
        totaal_km += km_dag
        omschrijvingen = [r.omschrijving or r.klant or "" for r in dag_ritten if (r.omschrijving or r.klant)]
        resultaat.append({
            "datum": dag.isoformat(),
            "datum_display": dag.strftime("%-d %B %Y"),
            "km_totaal": round(km_dag, 1),
            "type": "Office",
            "omschrijving": " / ".join(omschrijvingen) if omschrijvingen else "",
            "aantal_ritten": len(dag_ritten),
        })

    return {
        "jaar": jaar,
        "maand": maand,
        "maand_naam": MAANDEN_NL[maand],
        "totaal_km": round(totaal_km, 1),
        "aantal_dagen": len(resultaat),
        "invoer_per_dag": resultaat,
    }


@router.get("/excel/{jaar}/{maand}")
def export_excel(
    jaar: int,
    maand: int,
    db: Session = Depends(get_db),
    user: str = Depends(require_login),
):
    """Genereer een Excel-overzicht voor de maandelijkse SAP SuccessFactors invoer."""
    ritten = _ritten_voor_maand(db, jaar, maand)
    per_dag = _groepeer_per_dag(ritten)
    werknemer = _haal_instelling(db, "werknemer_naam", "Maurice Melotto")
    werknemer_nr = _haal_instelling(db, "werknemer_nummer", "")
    maand_naam = MAANDEN_NL[maand]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Km {maand_naam[:3]} {jaar}"

    # Kleuren
    BLAUW = "1F4E79"
    LICHTBLAUW = "BDD7EE"
    GRIJS = "F2F2F2"
    GROEN = "E2EFDA"
    WIT = "FFFFFF"

    def cel(rij, kolom, waarde=None, vet=False, kleur=None, uitlijning="left", formaat=None):
        c = ws.cell(row=rij, column=kolom, value=waarde)
        if vet:
            c.font = Font(bold=True, color=WIT if kleur == BLAUW else "000000")
        if kleur:
            c.fill = PatternFill(start_color=kleur, end_color=kleur, fill_type="solid")
        c.alignment = Alignment(horizontal=uitlijning, vertical="center")
        if formaat:
            c.number_format = formaat
        return c

    # Header-blok
    ws.merge_cells("A1:G1")
    c = ws.cell(row=1, column=1, value=f"Kilometerregistratie – {maand_naam.capitalize()} {jaar}")
    c.font = Font(bold=True, size=14, color=WIT)
    c.fill = PatternFill(start_color=BLAUW, end_color=BLAUW, fill_type="solid")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.cell(row=2, column=1, value="Werknemer:").font = Font(bold=True)
    ws.cell(row=2, column=2, value=werknemer)
    ws.cell(row=3, column=1, value="Personeelsnummer:").font = Font(bold=True)
    ws.cell(row=3, column=2, value=werknemer_nr)
    ws.cell(row=4, column=1, value="Systeem:").font = Font(bold=True)
    ws.cell(row=4, column=2, value="SAP SuccessFactors → Employment Information → Commuter Traffic")

    ws.row_dimensions[5].height = 10  # lege rij

    # Kolomheaders
    headers = ["Datum", "Dag", "Home/Office", "Kilometers (retour)", "Omschrijving / Klant", "# Ritten", "Status"]
    kolom_breedtes = [14, 12, 14, 22, 40, 10, 14]
    for i, (h, b) in enumerate(zip(headers, kolom_breedtes), 1):
        cel(6, i, h, vet=True, kleur=BLAUW, uitlijning="center")
        ws.column_dimensions[get_column_letter(i)].width = b
    ws.row_dimensions[6].height = 20

    # Dagerijen
    rij = 7
    totaal_km = 0.0
    for dag, dag_ritten in per_dag.items():
        km_dag = sum(r.afstand_km or 0 for r in dag_ritten)
        totaal_km += km_dag
        omschrijvingen = list({r.omschrijving or r.klant or "" for r in dag_ritten if (r.omschrijving or r.klant)})
        dag_naam = ["maandag", "dinsdag", "woensdag", "donderdag", "vrijdag", "zaterdag", "zondag"][dag.weekday()]
        achtergrond = GRIJS if rij % 2 == 0 else WIT

        cel(rij, 1, dag, kleur=achtergrond, uitlijning="center", formaat="DD-MM-YYYY")
        cel(rij, 2, dag_naam.capitalize(), kleur=achtergrond)
        cel(rij, 3, "Office", kleur=achtergrond, uitlijning="center")
        cel(rij, 4, round(km_dag, 1), kleur=achtergrond, uitlijning="center", formaat="0.0")
        cel(rij, 5, " / ".join(omschrijvingen), kleur=achtergrond)
        cel(rij, 6, len(dag_ritten), kleur=achtergrond, uitlijning="center")
        cel(rij, 7, "✓ Invoeren", kleur=achtergrond)
        ws.row_dimensions[rij].height = 18
        rij += 1

    # Totaalrij
    ws.merge_cells(f"A{rij}:C{rij}")
    cel(rij, 1, "TOTAAL", vet=True, kleur=LICHTBLAUW, uitlijning="right")
    cel(rij, 4, round(totaal_km, 1), vet=True, kleur=LICHTBLAUW, uitlijning="center", formaat="0.0")
    cel(rij, 5, f"{len(per_dag)} dagen", kleur=LICHTBLAUW)
    ws.row_dimensions[rij].height = 20
    rij += 2

    # Instructies
    ws.merge_cells(f"A{rij}:G{rij}")
    c = ws.cell(row=rij, column=1, value="📋  Invoer-instructie SAP SuccessFactors")
    c.font = Font(bold=True, size=11)
    c.fill = PatternFill(start_color=GROEN, end_color=GROEN, fill_type="solid")
    ws.row_dimensions[rij].height = 22
    rij += 1

    instructies = [
        "1. Open SAP SuccessFactors en klik op 'View my profile'",
        "2. Ga naar Employment Information → Commuter Traffic",
        "3. Klik op het potlood naast 'Home/Office Kilometer Info'",
        "4. Voer per dag (zie tabel hierboven) de datum in en kies 'Office'",
        "5. Vul het totaal aantal kilometers in (heen + terug = retour)",
        "6. Klik op 'Save' — herhaal voor elke reisdag",
        "⚠️  Deadline: uiterlijk de laatste dag van de maand invoeren!",
    ]
    for instr in instructies:
        ws.merge_cells(f"A{rij}:G{rij}")
        c = ws.cell(row=rij, column=1, value=instr)
        c.fill = PatternFill(start_color=GROEN, end_color=GROEN, fill_type="solid")
        if instr.startswith("⚠️"):
            c.font = Font(bold=True, color="C00000")
        ws.row_dimensions[rij].height = 16
        rij += 1

    # Borders op datatabellen
    dunne_rand = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    for r in range(6, rij):
        for col in range(1, 8):
            ws.cell(row=r, column=col).border = dunne_rand

    # Sla op in geheugen en stuur terug
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    bestandsnaam = f"km_successfactors_{jaar}_{maand:02d}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{bestandsnaam}"'},
    )
